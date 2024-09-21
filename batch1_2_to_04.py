import torch
import torch.nn as nn
import torch.nn.functional as F
import math
from torch.utils.tensorboard import SummaryWriter
import sys
import os
current_dir = os.path.dirname(__file__)
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
sys.path.append(parent_dir)
from lmmd import lmmd
from data_loader import data_loader
from openpyxl import *
from sklearn.decomposition import PCA
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')
import pandas as pd
from sklearn.manifold import TSNE

root_path = "./hsh/"
src1_name = "batch1"
src2_name = "batch2"
tgt_name = "batch4"
epochs = 600
log_interval = 10
batch_size = 12
lr = [0.00075, 0.0075]
l2_decay = 10e-3
momentum = 0.90
train_acc = []
test_acc = []

tra_loss = []
tst_loss = []
epo = []
font1 = {'family': 'Times New Roman',
         'weight': 'normal',
         'size': 18,
         }

no_cuda = False
seed = 8
cuda = not no_cuda and torch.cuda.is_available()
torch.manual_seed(seed)
if cuda:
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
device = torch.device("cuda" if cuda else "cpu")
kwargs = {'num_workers': 0, 'pin_memory': True} if cuda else {}

def plot_pca(data, labels, n_components=2, title='', save_path_image='', save_path_excel=''):
    pca = PCA(n_components=n_components)
    transformed_data = pca.fit_transform(data)
    df = pd.DataFrame(transformed_data, columns=[f'PC{i+1}' for i in range(n_components)])
    df['label'] = labels

    # Save PCA results to Excel
    df.to_excel(save_path_excel, index=False)

    # Plot PCA
    plt.figure(figsize=(8, 6))
    for label in df['label'].unique():
        subset = df[df['label'] == label]
        plt.scatter(subset['PC1'], subset['PC2'], label=f'Class {label}', alpha=0.5)

    plt.xlabel('Principal Component 1')
    plt.ylabel('Principal Component 2')
    plt.title(title)
    plt.legend()
    plt.savefig(save_path_image)
    plt.show()

def plot_tsne(data, labels, n_components=2, title='t-SNE (2D)', save_path_image=None, save_path_excel=None):
    tsne = TSNE(n_components=n_components)
    tsne_result = tsne.fit_transform(data)

    df = pd.DataFrame(tsne_result, columns=[f"dim{i + 1}" for i in range(n_components)])
    df['label'] = labels

    plt.figure(figsize=(8, 8))
    for label in df['label'].unique():
        indices = df['label'] == label
        plt.scatter(df.loc[indices, 'dim1'], df.loc[indices, 'dim2'], label=label)
    plt.title(title)
    plt.legend()
    if save_path_image:
        plt.savefig(save_path_image)
    if save_path_excel:
        df.to_excel(save_path_excel, index=False)
    plt.show()

class iAFF(nn.Module):
    '''
    多特征融合 iAFF
    '''
    def __init__(self, channels=1, r=1):
        super(iAFF, self).__init__()
        inter_channels = int(channels // r)

        # 局部注意力
        self.local_att = nn.Sequential(
            nn.Conv2d(channels, inter_channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(inter_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(inter_channels, channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(channels),
        )

        # 全局注意力
        self.global_att = nn.Sequential(
            nn.AdaptiveAvgPool2d(1),
            nn.Conv2d(channels, inter_channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(inter_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(inter_channels, channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(channels),
        )

        # 第二次局部注意力
        self.local_att2 = nn.Sequential(
            nn.Conv2d(channels, inter_channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(inter_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(inter_channels, channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(channels),
        )
        # 第二次全局注意力
        self.global_att2 = nn.Sequential(
            nn.AdaptiveAvgPool2d(1),
            nn.Conv2d(channels, inter_channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(inter_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(inter_channels, channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(channels),
        )

        self.sigmoid = nn.Sigmoid()

    def forward(self, x, residual):
        xa = x + residual
        xl = self.local_att(xa)
        xg = self.global_att(xa)
        xlg = xl + xg
        wei = self.sigmoid(xlg)
        xi = x * wei + residual * (1 - wei)

        xl2 = self.local_att2(xi)
        xg2 = self.global_att(xi)
        xlg2 = xl2 + xg2
        wei2 = self.sigmoid(xlg2)
        xo = x * wei2 + residual * (1 - wei2)
        return xo


class AFF(nn.Module):
    '''
    多特征融合 AFF
   '''
    def __init__(self, channels=1, r=1):
        super(AFF, self).__init__()
        inter_channels = int(channels // r)
        print(   inter_channels )
        self.local_att = nn.Sequential(
            nn.Conv2d(channels, inter_channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(inter_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(inter_channels, channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(channels),
        )
        # 全局注意力
        self.global_att = nn.Sequential(
            nn.AdaptiveAvgPool2d(1),
            nn.Conv2d(channels, inter_channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(inter_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(inter_channels, channels, kernel_size=1, stride=1, padding=0),
            nn.BatchNorm2d(channels),
        )
        self.sigmoid = nn.Sigmoid()

    def forward(self, x, residual):
        xa = x + residual
        xl = self.local_att(xa)
        xg = self.global_att(xa)
        xlg = xl + xg
        wei = self.sigmoid(xlg)

        xo = x * wei + residual * (1 - wei)
        return xo

class PositionalEncoding(nn.Module):

    def __init__(self, d_model, max_len=500):
        super(PositionalEncoding, self).__init__()
        pe = torch.zeros(max_len, d_model)
        position = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        div_term = torch.exp(torch.arange(0, d_model, 2).float() * (-math.log(10000.0) / d_model))
        pe[:, 0::2] = torch.sin(position * div_term)
        pe[:, 1::2] = torch.cos(position * div_term)
        pe = pe.unsqueeze(0).transpose(0, 1)

        # pe.requires_grad = False
        self.register_buffer('pe', pe)

    def forward(self, x):
        return x + self.pe[:x.size(0), :]

class TransAm(nn.Module):
    def __init__(self, feature_size=128, num_layers=4, dropout=0.2):
        super(TransAm, self).__init__()
        self.model_type = 'Transformer'
        self.src_mask = None
        self.pos_encoder = PositionalEncoding(feature_size)
        self.encoder_layer = nn.TransformerEncoderLayer(d_model=feature_size, nhead=64, dropout=dropout)
        self.transformer_encoder = nn.TransformerEncoder(self.encoder_layer, num_layers=num_layers)
        self.decoder_1 = nn.Linear(feature_size, 128)

    def forward(self, src):
        src = self.pos_encoder(src)
        output = self.transformer_encoder(src)
        output = self.decoder_1(output)
        return output

class Res_TransAm1(nn.Module):
    def __init__(self, feature_size=128, num_layers=4, dropout=0.2):
        super(Res_TransAm1, self).__init__()
        self.model_type = 'Transformer'
        self.src_mask = None
        self.pos_encoder = PositionalEncoding(feature_size)
        self.encoder_layer = nn.TransformerEncoderLayer(d_model=feature_size, nhead=64, dropout=dropout)
        self.transformer_encoder = nn.TransformerEncoder(self.encoder_layer, num_layers=num_layers)
        self.decoder_1 = nn.Linear(feature_size, 64)

    def forward(self, src):
        src = self.pos_encoder(src)
        output = self.transformer_encoder(src)
        output = self.decoder_1(output)
        return output

class Res_TransAm2(nn.Module):
    def __init__(self, feature_size=128, num_layers=4, dropout=0.2):
        super(Res_TransAm2, self).__init__()
        self.model_type = 'Transformer'
        self.src_mask = None
        self.pos_encoder = PositionalEncoding(feature_size)
        self.encoder_layer = nn.TransformerEncoderLayer(d_model=feature_size, nhead=64, dropout=dropout)
        self.transformer_encoder = nn.TransformerEncoder(self.encoder_layer, num_layers=num_layers)
        self.decoder_1 = nn.Linear(feature_size, 64)

    def forward(self, src):
        src = self.pos_encoder(src)
        output = self.transformer_encoder(src)
        output = self.decoder_1(output)
        return output


class Net(nn.Module):
    def __init__(self, num_classes=6):
        super(Net, self).__init__()
        self.fusion = iAFF().to(device)
        self.dense = TransAm().to(device)
        self.cls_fc_son1_2 = nn.Conv1d(128, 64, 3, padding=1)
        self.cls_fc_son2_2 = nn.Conv1d(128, 64, 3, padding=1)
#----------------------------------------------------------------私有特征跳接层，使用的是全连接神经网络。对应的是模型图中的G1和G2部分--------------------------------------------------
        # self.res1 = nn.Linear(128, 50)
        # self.res2 = nn.Linear(128, 50)
        self.res1 = Res_TransAm1().to(device)
        self.res2 = Res_TransAm2().to(device)

#-----------------------------------------------------------------接在私有特征叠加融合后进行降维与分类训练。对应的是模型图中的C1和C2部分-------------------------------------------------
        self.cls_fc1_1 = nn.Linear(64, 25)
        self.cls_fc2_1 = nn.Linear(64, 25)
        self.cls_fc1_2 = nn.Linear(25, num_classes)
        self.cls_fc2_2 = nn.Linear(25, num_classes)

#------------------------------------------------------------域对齐损失函数，本模型使用两次，一次对跳接层的对齐，一次为对私有域融合后的对齐-----------------------------------------------
        self.lmmd_loss = lmmd.LMMD_loss(class_num=num_classes)
        # self.mmd_loss = MMD.MMDLoss()

    def forward(self, source, target, s_label, mark=1):
# ---------------------------------------------------------------------------源域1和目标域进行计算与对齐-----------------------------------------------------------------------
        if mark == 1:
            source = source.unsqueeze(0)                  # 扩充一个维度 (1, 24, 128)
            s_fea = (self.dense(source))                  # 源领域进入common features extractor (1, 24, 64)
            source = source.squeeze()                     # (24, 128)
            s_fea = s_fea.permute(1, 2, 0)                # (24, 128, 1)
            # print(s_fea.shape)                            #
            s_fea2 = self.cls_fc_son1_2(s_fea)            # 一维卷积 (24, 50, 1)     (batch_size, seq_length, embedding_dim)
            # print(s_fea2.shape)
            s_fea2 = s_fea2.permute(2, 0, 1)              #  （1, 24, 50）
            s_fea2 = s_fea2.squeeze(0)                    # （24, 50）
            s_fea2 = s_fea2.reshape(-1, 8, 8)              # (24, 5, 10)
            s_fea2 = s_fea2.unsqueeze(1)                    # (24, 1, 5, 10)

            source = source.unsqueeze(0)
            s_res = self.res1(source)                     # (1, 24, 50)
            s_res_l = s_res.squeeze()                       # (24, 50)
            s_res = s_res_l.reshape(-1, 8, 8)              # (24, 5, 10)
            s_res = s_res.unsqueeze(1)                    # (24, 1, 5, 10)
            s_fusion = self.fusion(s_fea2, s_res)
            s_fusion1 = s_fusion.squeeze(1).reshape(-1, 64)
            s_need_pred = self.cls_fc1_1(s_fusion1)
            s_pred_1 = F.relu(s_need_pred)                # relu activation function
            s_pred = self.cls_fc1_2(s_pred_1)             #  (24, 6)

            target = target.unsqueeze(0)
            t_fea = (self.dense(target))
            target = target.squeeze()
            t_fea = t_fea.permute(1, 2, 0)
            t_fea2 = self.cls_fc_son1_2(t_fea)
            t_fea2 = t_fea2.permute(2, 0, 1)
            t_fea2 = t_fea2.squeeze(0)
            t_fea2 = t_fea2.reshape(-1, 8, 8)              # (24, 5, 10)
            t_fea2 = t_fea2.unsqueeze(1)                    # (24, 1, 5, 10)
            target = target.unsqueeze(0)
            t_res = self.res1(target)
            t_res_l = t_res.squeeze()
            t_res = t_res_l.reshape(-1, 8, 8)              # (24, 5, 10)
            t_res = t_res.unsqueeze(1)                    # (24, 1, 5, 10)
            t_fusion = self.fusion(t_fea2, t_res)
            t_fusion = t_fusion.squeeze(1).reshape(-1, 64)
            t_need_pred = self.cls_fc1_1(t_fusion)

            target = target.squeeze()
            t_label_1 = F.relu(t_need_pred)
            t_label = self.cls_fc1_2(t_label_1)
            # loss = self.lmmd_loss.get_loss(s_fea2, t_fea2, s_label, torch.nn.functional.softmax(t_label, dim=1))
            loss = self.lmmd_loss.get_loss(s_need_pred, t_need_pred, s_label, torch.nn.functional.softmax(t_label, dim=1))
            loss += self.lmmd_loss.get_loss(s_res_l, t_res_l, s_label, torch.nn.functional.softmax(t_label, dim=1))
            # loss += self.mmd_loss(s_res, t_res)

            target = target.unsqueeze(0)
            t2_fea = (self.dense(target))
            target = target.squeeze(0)
            t2_fea = t2_fea.permute(1, 2, 0)
            t2_fea2 = self.cls_fc_son2_2(t2_fea)
            t2_fea2 = t2_fea2.permute(2, 0, 1)
            t2_fea2 = t2_fea2.squeeze(0)
            t2_fea2 = t2_fea2.reshape(-1, 8, 8)              # (24, 5, 10)
            t2_fea2 = t2_fea2.unsqueeze(1)                    # (24, 1, 5, 10)
            target = target.unsqueeze(0)
            t_res = self.res2(target)
            t_res = t_res.squeeze()
            t_res = t_res.reshape(-1, 8, 8)              # (24, 5, 10)
            t_res = t_res.unsqueeze(1)                    # (24, 1, 5, 10)
            t_fusion = self.fusion(t2_fea2, t_res)
            t_fusion = t_fusion.squeeze(1).reshape(-1, 64)
            t2_label_1 = F.relu(self.cls_fc2_1(t_fusion))
            t2_label = self.cls_fc2_2(t2_label_1)

            l1_loss = torch.abs(
                torch.nn.functional.softmax(t_label, dim=1) - torch.nn.functional.softmax(t2_label, dim=1))
            l1_loss = torch.mean(l1_loss)

            cls_loss = F.nll_loss(F.log_softmax(s_pred, dim=1), s_label)

            return cls_loss, loss, l1_loss, s_fusion1

# ---------------------------------------------------------------------------源域2和目标域进行计算与对齐-----------------------------------------------------------------------
        if mark == 2:
            source = source.unsqueeze(0)    # 扩充一个维度
            s_fea = (self.dense(source))    # 进common feature extractor
            source = source.squeeze()       # 压缩一个维度
            s_fea = s_fea.permute(1, 2, 0)  # 维度交换
            s_fea2 = self.cls_fc_son2_2(s_fea)
            s_fea2 = s_fea2.permute(2, 0, 1)
            s_fea2 = s_fea2.squeeze(0)
            s_fea2 = s_fea2.reshape(-1, 8, 8)  # (24, 5, 10)
            s_fea2 = s_fea2.unsqueeze(1)        # (24, 1, 5, 10)
            source = source.unsqueeze(0)
            s_res = self.res2(source)
            s_res_l = s_res.squeeze()
            s_res = s_res_l.reshape(-1, 8, 8)  # (24, 5, 10)
            s_res = s_res.unsqueeze(1)
            s_fusion = self.fusion(s_fea2, s_res)  #
            s_fusion2 = s_fusion.squeeze(1).reshape(-1, 64)
            s_need_pred = self.cls_fc2_1(s_fusion2)
            s_pred_1 = F.relu(s_need_pred)
            s_pred = self.cls_fc2_2(s_pred_1)

            target = target.unsqueeze(0)
            t_fea = (self.dense(target))
            target = target.squeeze()
            t_fea = t_fea.permute(1, 2, 0)
            t_fea2 = self.cls_fc_son2_2(t_fea)
            t_fea2 = t_fea2.permute(2, 0, 1)
            t_fea2 = t_fea2.squeeze(0)
            t_fea2 = t_fea2.reshape(-1, 8, 8)  # (24, 5, 10)
            t_fea2 = t_fea2.unsqueeze(1)        # (24, 1, 5, 10)
            target = target.unsqueeze(0)
            t_res = self.res2(target)
            t_res_l = t_res.squeeze()
            t_res = t_res_l.reshape(-1, 8, 8)  # (24, 5, 10)
            t_res = t_res.unsqueeze(1)        # (24, 1, 5, 10)
            t_fusion = self.fusion(t_fea2, t_res)  #
            t_fusion = t_fusion.squeeze(1).reshape(-1, 64)
            t_need_pred = self.cls_fc2_1(t_fusion)
            target = target.squeeze()
            t_label_1 = F.relu(t_need_pred)
            t_label = self.cls_fc2_2(t_label_1)
            loss = self.lmmd_loss.get_loss(s_need_pred, t_need_pred, s_label, torch.nn.functional.softmax(t_label, dim=1))
            loss +=  self.lmmd_loss.get_loss(s_res_l, t_res_l, s_label, torch.nn.functional.softmax(t_label, dim=1))


            target = target.unsqueeze(0)
            t2_fea = (self.dense(target))
            target = target.squeeze()
            t2_fea = t2_fea.permute(1, 2, 0)
            t2_fea2 = self.cls_fc_son1_2(t2_fea)
            t2_fea2 = t2_fea2.permute(2, 0, 1)
            t2_fea2 = t2_fea2.squeeze(0)
            t2_fea2 = t2_fea2.reshape(-1, 8, 8)  # (24, 5, 10)
            t2_fea2 = t2_fea2.unsqueeze(1)        # (24, 1, 5, 10)
            target = target.unsqueeze(0)
            t_res = self.res2(target)
            t_res = t_res.squeeze()
            t_res = t_res.reshape(-1, 8, 8)  # (24, 5, 10)
            t_res = t_res.unsqueeze(1)        # (24, 1, 5, 10)
            t_fusion = self.fusion(t2_fea2, t_res)  #
            t_fusion = t_fusion.squeeze(1).reshape(-1, 64)
            t2_label_1 = F.relu(self.cls_fc1_1(t_fusion))
            # t2_label_1 = F.relu(self.cls_fc1_1(torch.cat([t2_fea2, self.res2(target)], dim=1)))

            # torch.cat([s_fea2, s_res], dim=1)
            t2_label = self.cls_fc1_2(t2_label_1)

            l1_loss = torch.abs(
                torch.nn.functional.softmax(t_label, dim=1) - torch.nn.functional.softmax(t2_label, dim=1))
            l1_loss = torch.mean(l1_loss)

            cls_loss = F.nll_loss(F.log_softmax(s_pred, dim=1), s_label)

            return cls_loss, loss, l1_loss, s_fusion2
        else:
            return 0

    def predict(self, source):

        source = source.unsqueeze(0)
        s_fea = (self.dense(source))
        source = source.squeeze()
        s_fea = s_fea.permute(1, 2, 0)
        s_fea2 = self.cls_fc_son1_2(s_fea)
        s_fea2 = s_fea2.permute(2, 0, 1)
        s_fea2 = s_fea2.squeeze(0)
        s_fea2 = s_fea2.reshape(-1, 8, 8)  # (24, 5, 10)
        s_fea2 = s_fea2.unsqueeze(1)       # (24, 1, 5, 10)
        source = source.unsqueeze(0)
        s_res = self.res1(source)          # (24, 50)
        s_res = s_res.squeeze()
        s_res = s_res.reshape(-1, 8, 8)    # (24, 5, 10)
        s_res = s_res.unsqueeze(1)         # (24, 1, 5, 10)
        s_fusion = self.fusion(s_fea2, s_res)  #
        t_fusion1 = s_fusion.squeeze(1).reshape(-1, 64)
        s_pred1_1 = F.relu(self.cls_fc1_1(t_fusion1))
        source = source.squeeze()
        s_pred1_1  = s_pred1_1.squeeze()
        s_pred1 = self.cls_fc1_2(s_pred1_1)

        source = source.unsqueeze(0)
        s2_fea = (self.dense(source))
        source = source.squeeze()
        s2_fea = s2_fea.permute(1, 2, 0)
        s2_fea2 = self.cls_fc_son2_2(s2_fea)
        s2_fea2 = s2_fea2.permute(2, 0, 1)
        s2_fea2 = s2_fea2.squeeze(0)
        s2_fea2 = s2_fea2.reshape(-1, 8, 8)  # (24, 5, 10)
        s2_fea2 = s2_fea2.unsqueeze(1)  # (24, 1, 5, 10)
        source = source.unsqueeze(0)
        s_res2 = self.res2(source)
        s_res2 = s_res2.squeeze()
        s_res2 = s_res2.reshape(-1, 8, 8)  # (24, 5, 10)
        s_res2 = s_res2.unsqueeze(1)  # (24, 1, 5, 10)
        s_fusion = self.fusion(s2_fea2, s_res2)  #
        t_fusion2 = s_fusion.squeeze(1).reshape(-1, 64)
        s2_pred2_1 = F.relu(self.cls_fc2_1(t_fusion2))
        source = source.squeeze()
        s_pred2 = self.cls_fc2_2(s2_pred2_1)

        return s_pred1, s_pred2, t_fusion1, t_fusion2

def train_epoch(epoch, model, dataloaders, optimizer, writer):
    labels_fusion1 = []
    labels_fusion2 = []
    sre_fusion1 = None
    sre_fusion2 = None
    model.train()
    source1_loader, source2_loader, target_train_loader, _ = dataloaders
    source1_iter = iter(source1_loader)
    source2_iter = iter(source2_loader)
    target_iter = iter(target_train_loader)
    num_iter = min(len(source1_loader), len(source2_loader))
    for i in range(1, num_iter):
        try:
            source_data, source_label = next(source1_iter)
        except Exception as err:
            source1_iter = iter(source1_loader)
            source_data, source_label = next(source1_iter)
        try:
#            target_data, __ = target_iter.next()
             target_data, __ = next(target_iter)
        except Exception as err:
            target_iter = iter(target_train_loader)
#           target_data, __ = target_iter.next()
            target_data, __ = next(target_iter)
        if cuda:
            source_data, source_label = source_data.type(torch.FloatTensor).cuda(), source_label.type(
                torch.LongTensor).cuda()
            target_data = target_data.type(torch.FloatTensor).cuda()    # 浮点型的Tensor，传来参数可以是一个列表

        optimizer.zero_grad()                                      # 先将梯度归零
        cls_loss, lmmd_loss, l1_loss, s_fusion1= model(                      # 获取loss
            source_data, target_data, source_label - 1, mark=1)    # source_data: [24,128]
        labels_fusion1.append(source_label)
        lambd = 2 / (1 + math.exp(-0.1 * (epoch) / epochs)) - 1
        loss = cls_loss +  lambd * (lmmd_loss + l1_loss)

        writer.add_scalar('TRAINwithMFSAN/Loss', loss.item(), i + epoch * num_iter)
        loss.backward()                                                  # 反向传播计算得到每个参数的梯度值
        optimizer.step()                                                 # 根据梯度更新网络参数

        if i % log_interval == 0:
            print(
                f'source1-Epoch: [{epoch:2d}], Loss: {loss.item():.4f}, cls_Loss: {cls_loss.item():.4f}, loss_lmmd: {lmmd_loss.item():.4f}, l1_loss: {l1_loss.item():.4f}')

        try:
#            source_data, source_label = source2_iter.next()
            source_data, source_label = next(source2_iter)
        except Exception as err:
            source2_iter = iter(source2_loader)
#            source_data, source_label = source2_iter.next()
            source_data, source_label = next(source2_iter)
        try:
#            target_data, __ = target_iter.next()
             target_data, __ = \
                 (target_iter)
        except Exception as err:
            target_iter = iter(target_train_loader)
#           target_data, __ = target_iter.next()
            target_data, __ = next(target_iter)
        if cuda:
            source_data, source_label = source_data.type(torch.FloatTensor).cuda(), source_label.type(
                torch.LongTensor).cuda()
            target_data = target_data.type(torch.FloatTensor).cuda()
#            target_data = target_data.type(torch.FloatTensor).cuda()

        optimizer.zero_grad()
        cls_loss, lmmd_loss, l1_loss, s_fusion2 = model(
            source_data, target_data, source_label - 1, mark=2)
        labels_fusion2.append(source_label)
        lambd = 2 / (1 + math.exp(-0.1* (epoch) / epochs)) - 1
        loss = cls_loss +   lambd * (lmmd_loss +  l1_loss)
#         loss = cls_loss + (lmmd_loss + l1_loss)
        # loss=cls_loss
        writer.add_scalar('TRAINwithMFSAN/Loss', loss.item(), i + epoch * num_iter)

        loss.backward()
        optimizer.step()

        if i % log_interval == 0:
            print(
                f'source2-Epoch: [{epoch:2d}], Loss: {loss.item():.4f}, cls_Loss: {cls_loss.item():.4f}, lmmd_loss: {lmmd_loss.item():.4f}, l1_loss: {l1_loss.item():.4f}')

        if sre_fusion1 is None:
            sre_fusion1 = s_fusion1
        else:
            sre_fusion1 = torch.cat((sre_fusion1, s_fusion1), dim=0)

        if sre_fusion2 is None:
            sre_fusion2 = s_fusion2
        else:
            sre_fusion2 = torch.cat((sre_fusion2, s_fusion2), dim=0)
    labels_fusion1 = torch.cat(labels_fusion1, dim=0) if labels_fusion1 else torch.empty(0)
    labels_fusion2 = torch.cat(labels_fusion2, dim=0) if labels_fusion2 else torch.empty(0)

    return sre_fusion1, sre_fusion2, labels_fusion1, labels_fusion2

def load_data(root_path, src1, src2, tar, batch_size):
    loader_src1 = data_loader.load_dataset(root_path+'train_new/', src1, batch_size, kwargs)
    loader_src2 = data_loader.load_dataset(root_path+'train_new/', src2, batch_size, kwargs)
    loader_tar = data_loader.load_dataset(root_path+'train_new/', tar, batch_size, kwargs)
    loader_tar_test = data_loader.load_dataset(root_path+'train_new/', tar, batch_size, kwargs)
    return loader_src1, loader_src2, loader_tar, loader_tar_test

def train(model, dataloader):
    model.eval()
    train_loss = 0
    correct = 0
    correct3 = 0
    batch_num = 0
    with torch.no_grad():                      # 当前计算不需要反向传播
        for data, target in dataloader:
            batch_num += 1
            if cuda:
                data, target = data.type(torch.FloatTensor).cuda(), target.type(torch.LongTensor).cuda()
            pred1, pred2, _, _ = model.predict(data)
            pred = 0.5 * pred1 + 0.5 * pred2
            train_loss += F.nll_loss(F.log_softmax(pred, dim=1), target - 1).item()  # negative log likelihood loss
            pred = pred.data.max(1)[1]
            correct += pred.eq((target - 1).data.view_as(pred)).cpu().sum()
        train_loss /= (batch_num * batch_size) #  test_loss = test_loss / (batch_num * batch_size)
        acc = correct/(batch_num * batch_size)
        train_acc.append(acc)
        tra_loss.append(train_loss)

    return train_acc, correct, batch_num

def tst(model, dataloader):
    targets = []
    tgt_fusion1 = []
    tgt_fusion2 = []
    model.eval()                               #  model.eval()，作用等同于self.train(False) tchNorm层，dropout层等用于优化训练而添加的网络层会被关闭
    test_loss = 0
    correct = 0
    correct3 = 0
    batch_num = 0
    with torch.no_grad():                      # 当前计算不需要反向传播
        for data, target in dataloader:
            batch_num += 1
            if cuda:
                data, target = data.type(torch.FloatTensor).cuda(), target.type(torch.LongTensor).cuda()
            pred1, pred2, t_fusion1, t_fusion2 = model.predict(data)
            pred = 0.5 * pred1 + 0.5 * pred2
            test_loss += F.nll_loss(F.log_softmax(pred, dim=1), target - 1).item()  # negative log likelihood loss
            pred = pred.data.max(1)[1]
            correct += pred.eq((target - 1).data.view_as(pred)).cpu().sum()
            tgt_fusion1.append(t_fusion1.view(t_fusion1.size(0), -1))  # Flatten or reshape if necessary
            tgt_fusion2.append(t_fusion2.view(t_fusion2.size(0), -1))  # Flatten or reshape if necessary
            targets.append(target)
        test_loss /= (batch_num * batch_size) #  test_loss = test_loss / (batch_num * batch_size)
        acc = correct/(batch_num * batch_size)
        test_acc.append(acc)
        tst_loss.append(test_loss)
        print(
            f'Average loss: {test_loss:.4f}, Test_Accuracy: {correct}/{(batch_num * batch_size)} ({100. * correct / (batch_num * batch_size):.2f}%)')
        tgt_fusion1 = torch.cat(tgt_fusion1, dim=0)
        tgt_fusion2 = torch.cat(tgt_fusion2, dim=0)
        targets = torch.cat(targets, dim=0)
    return test_acc, correct, batch_num, tgt_fusion1, tgt_fusion2, targets

def main():
    dataloaders = load_data(root_path, src1_name, src2_name, tgt_name, batch_size)
    model = Net().cuda()
    correct = 0
    stop = 0

    optimizer = torch.optim.SGD([
        {'params': model.dense.parameters()},
        # {'params': model.dense1.parameters()},
        # {'params': model.dense2.parameters()},
        # {'params': model.cls_fc_son1_1.parameters()},
        {'params': model.cls_fc_son1_2.parameters()},
        {'params': model.res1.parameters()},
        {'params': model.res2.parameters()},
        # {'params': model.cls_fc_son2_1.parameters()},
        {'params': model.cls_fc_son2_2.parameters()},
        {'params': model.cls_fc1_1.parameters(), 'lr': lr[0]},
        {'params': model.cls_fc1_2.parameters(), 'lr': lr[0]},
        {'params': model.cls_fc2_1.parameters(), 'lr': lr[0]},
        {'params': model.cls_fc2_2.parameters(), 'lr': lr[0]},
    ], lr=lr[0], momentum=momentum, weight_decay=l2_decay)
    writer = SummaryWriter(log_dir='./runs/train_loss', flush_secs=20)
    writer1 = SummaryWriter(log_dir='./runs/test_acc', flush_secs=20)

    for epoch in range(1, epochs + 1):
        epo.append(epoch)
        # stop += 1
        optimizer.param_groups[0]['lr'] = lr[0] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.75)     # 设置学习率
        optimizer.param_groups[1]['lr'] = lr[0] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.75)
        optimizer.param_groups[2]['lr'] = lr[0] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.75)
        optimizer.param_groups[3]['lr'] = lr[0] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.5)
        optimizer.param_groups[4]['lr'] = lr[0] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.5)
        # optimizer.param_groups[5]['lr'] = lr[1] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.5)
        # optimizer.param_groups[6]['lr'] = lr[1] / math.pow((1 + 10 * (epoch - 1) / epochs), 1.5)
        s_fusion1, s_fusion2, labels_fusion1, labels_fusion2 = train_epoch(epoch, model, dataloaders, optimizer, writer)            #
        train_acc, train_correct, num = train(model, dataloaders[2])
        test_acc, t_correct, num, t_fusion1, t_fusion2, targets = tst(model, dataloaders[-1])

        plt.figure(1)
        plt.plot(epo, train_acc, color='green', linewidth=1.5, label='train accuracy')
        plt.plot(epo, test_acc, color='red', linewidth=1.5, label='test accuracy')
        plt.xlabel("epoch", font1, color='k')  # 指定X坐标轴的标签，并设置标签字体大小
        plt.ylabel("accuracy", font1, color='k')  # 指定Y坐标轴的标签，并设置标签字体大小
        #       plt.title('test accuracy', font1, color='k')
        plt.legend(loc='lower right', prop=font1)

        plt.figure(2)
        plt.plot(epo, tra_loss, color='green', linewidth=1.5, label='train loss')
        plt.plot(epo, tst_loss, color='red', linewidth=1.5, label='test loss')
        plt.xlabel("epoch", font1, color='k')  # 指定X坐标轴的标签，并设置标签字体大小
        plt.ylabel("loss", font1, color='k')  # 指定Y坐标轴的标签，并设置标签字体大小
        plt.legend(loc='upper right', prop=font1)
        plt.show()

        wb = load_workbook(r"data_01/acc-loss-12-04.xlsx")
        ws = wb['Sheet1']
        for col in range(len(test_acc)):
            c = col + 2
            ws.cell(row=c, column=1).value = float(test_acc[col])

        for col in range(len(tst_loss)):
            c = col + 2
            ws.cell(row=c, column=2).value = float(tst_loss[col])
        wb.save(r"data_01/acc-loss-12-04.xlsx")

        writer1.add_scalar('TESTwithMFSAN/%', 100. * t_correct / len(dataloaders[-1].dataset), epoch)
        if t_correct > correct:
            correct = t_correct
            stop = 0
            torch.save(model.state_dict(), 'trained_models/' + tgt_name + '.pt')
        print(
            f'{src1_name}，{src2_name}-{tgt_name}: max correct: {correct} max accuracy: {100. * correct / (num * batch_size):.2f}%\n')

        if stop >= 140:
            print(
                f'Final test acc: {100. * correct / (num * batch_size):.2f}%')
            break

    torch.save(model.state_dict(), 'trained_models/' + tgt_name + '.pt')

    t_fusion1_cpu = t_fusion1.cpu().numpy()
    t_fusion2_cpu = t_fusion2.cpu().numpy()
    targets_cpu = targets.cpu().numpy()

    s_fusion1_cpu = s_fusion1.detach().cpu().numpy()
    s_fusion2_cpu = s_fusion2.detach().cpu().numpy()
    labels_fusion1_cpu = labels_fusion1.detach().cpu().numpy()
    labels_fusion2_cpu = labels_fusion2.detach().cpu().numpy()

    plot_pca(t_fusion1_cpu, targets_cpu, n_components=2, title='t_fusion1 PCA (2D)',
             save_path_image=f'pca_12_04/t_fusion1_pca_2d_epoch_{epoch}.png',
             save_path_excel=f'pca_12_04/t_fusion1_pca_2d_epoch_{epoch}.xlsx')
    plot_pca(t_fusion2_cpu, targets_cpu, n_components=2, title='t_fusion2 PCA (2D)',
             save_path_image=f'pca_12_04/t_fusion2_pca_2d_epoch_{epoch}.png',
             save_path_excel=f'pca_12_04/t_fusion2_pca_2d_epoch_{epoch}.xlsx')
    plot_pca(s_fusion1_cpu, labels_fusion1_cpu, n_components=2, title='s_fusion1 PCA (2D)',
             save_path_image = f'pca_12_04/s_fusion1_pca_2d_epoch_{epoch}.png',
             save_path_excel = f'pca_12_04/s_fusion1_pca_2d_epoch_{epoch}.xlsx')
    plot_pca(s_fusion2_cpu,  labels_fusion2_cpu, n_components=2, title='s_fusion2 PCA (2D)',
             save_path_image = f'pca_12_04/s_fusion2_pca_2d_epoch_{epoch}.png',
             save_path_excel = f'pca_12_04/s_fusion2_pca_2d_epoch_{epoch}.xlsx')
    print(f"Epoch {epoch}: PCA plots generated and saved.")

    # Plot using t-SNE
    plot_tsne(t_fusion1_cpu, targets_cpu, n_components=2, title='t_fusion1 t-SNE (2D)',
               save_path_image=f'pca_12_04/t_fusion1_tsne_2d_epoch_{epoch}.png',
               save_path_excel=f'pca_12_04/t_fusion1_tsne_2d_epoch_{epoch}.xlsx')
    plot_tsne(t_fusion2_cpu, targets_cpu, n_components=2, title='t_fusion2 t-SNE (2D)',
               save_path_image=f'pca_12_04/t_fusion2_tsne_2d_epoch_{epoch}.png',
               save_path_excel=f'pca_12_04/t_fusion2_tsne_2d_epoch_{epoch}.xlsx')
    plot_tsne(s_fusion1_cpu, labels_fusion1_cpu, n_components=2, title='s_fusion1 t-SNE (2D)',
               save_path_image=f'pca_12_04/s_fusion1_tsne_2d_epoch_{epoch}.png',
               save_path_excel=f'pca_12_04/s_fusion1_tsne_2d_epoch_{epoch}.xlsx')
    plot_tsne(s_fusion2_cpu, labels_fusion2_cpu, n_components=2, title='s_fusion2 t-SNE (2D)',
               save_path_image=f'pca_12_04/s_fusion2_tsne_2d_epoch_{epoch}.png',
               save_path_excel=f'pca_12_04/s_fusion2_tsne_2d_epoch_{epoch}.xlsx')
    print(f"Epoch {epoch}: t-SNE plots generated and saved.")

def print_hi(name):
    print(f'Hi, {name}')


if __name__ == '__main__':
    main()

