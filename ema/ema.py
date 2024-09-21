import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')
import pandas as pd
from openpyxl import *
plt.rcParams['font.sans-serif'] = ['arial']  # 设置字体为中文宋体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
alpha1 = 0.1
alpha2 = 0.01
alpha3 = 0.001
df = pd.read_csv('180ppmH2_CO2.txt', sep='\t', header=None)
col1 = df.iloc[:, 6]

def exponential_moving_average(data, alpha):
    ema = []
    ema.append(0)
    for i in range(1, len(data)):
        ema.append(alpha * (data[i] - data[i-1])+ (1 - alpha) * ema[i - 1])
    return ema

ema1 = exponential_moving_average(col1, alpha1)
ema2 = exponential_moving_average(col1, alpha2)
ema3 = exponential_moving_average(col1, alpha3)

plt.figure(1)
plt.plot( col1, label='raw signal')
plt.xlabel("times / (s)",  color='k')
plt.ylabel("voltage / (V)",   color='k')

wb = load_workbook(r"ema_data.xlsx")
ws = wb['Sheet1']
for col in range(len(col1)):
    c = col + 2
    ws.cell(row=c, column=1).value = float(col1[col])

plt.show()

plt.figure(2)
plt.plot(range(len(col1)), ema1, label='ema')
for col in range(len(ema1)):
    c = col + 2
    ws.cell(row=c, column=2).value = float(ema1[col])

plt.figure(3)
plt.plot(range(len(col1)), ema2, label='ema')
for col in range(len(ema1)):
    c = col + 2
    ws.cell(row=c, column=3).value = float(ema2[col])

plt.figure(4)
plt.plot(range(len(col1)), ema3, label='ema')
for col in range(len(ema1)):
    c = col + 2
    ws.cell(row=c, column=4).value = float(ema3[col])
wb.save(r"ema_data.xlsx")
plt.legend()
plt.show()


